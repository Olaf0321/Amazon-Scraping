import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import time
import openpyxl
import os
from scrap import scrap_info
from datetime import datetime
import pytz

def get_current_time():
    # Get the Japan timezone
    japan_tz = pytz.timezone('Asia/Tokyo')

    # Get the current time in Japan timezone
    japan_time = datetime.now(japan_tz)

    # Print the current time in Japan
    return japan_time.strftime('%Y-%m-%d %H-%M-%S')
class ExcelProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("Amazon 商品のスクレイピング")
        self.root.geometry("800x500")
        self.root.configure(bg="#F0F0F0")
        self.root.resizable(False, False)
        
        self.input_path = tk.StringVar()
        self.output_folder = tk.StringVar()
        self.running = False
        
        style = ttk.Style()
        style.theme_use("clam")
        
        # Button Styling with Hover Effects
        style.configure("TButton", font=("Helvetica Neue", 12), padding=10, background="#4CAF50", foreground="white", relief="flat")
        style.map("TButton", background=[("active", "#45a049"), ("!disabled", "#388E3C")])
        style.configure("TLabel", font=("Helvetica Neue", 12), background="#FFFFFF", foreground="#333333")
        style.configure("TEntry", font=("Helvetica Neue", 12), padding=10, fieldbackground="#F4F4F4", foreground="#333333", relief="flat")
        style.configure("TFrame", background="#FFFFFF")
        style.configure("TProgressbar", thickness=15, troughcolor="#E0E0E0", background="#4CAF50", borderwidth=1)

        # Main Frame with Gradient Background
        main_frame = ttk.Frame(root)
        main_frame.pack(pady=30, padx=30, fill=tk.BOTH, expand=True)

        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=15)
        ttk.Label(input_frame, text="入力ファイル:").pack(side=tk.LEFT, padx=10)
        input_entry = ttk.Entry(input_frame, textvariable=self.input_path, width=60, state="readonly")
        input_entry.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
        self.input_browse_button = ttk.Button(input_frame, text="ブラウズ", command=self.select_input_path)
        self.input_browse_button.pack(side=tk.LEFT, padx=10)

        output_frame = ttk.Frame(main_frame)
        output_frame.pack(fill=tk.X, pady=15)
        ttk.Label(output_frame, text="出力フォルダ:").pack(side=tk.LEFT, padx=10)
        output_entry = ttk.Entry(output_frame, textvariable=self.output_folder, width=60, state="readonly")
        output_entry.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
        self.output_browse_button = ttk.Button(output_frame, text="ブラウズ", command=self.select_output_folder)
        self.output_browse_button.pack(side=tk.LEFT, padx=10)
        
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=30)
        self.start_button = ttk.Button(button_frame, text="実行", command=self.start_processing, width=18)
        self.start_button.pack(side=tk.LEFT, padx=15)
        self.stop_button = ttk.Button(button_frame, text="停止", command=self.stop_processing, state=tk.DISABLED, width=18)
        self.stop_button.pack(side=tk.LEFT, padx=15)

        self.status_label = ttk.Label(main_frame, text="状態: 待機中", font=("Helvetica Neue", 14))
        self.status_label.pack(pady=20)

        self.progress = ttk.Progressbar(main_frame, length=650, mode="determinate")
        self.progress.pack(pady=15, fill=tk.X)
    
    def select_input_path(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if path:
            self.input_path.set(path)
    
    def select_output_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.output_folder.set(folder)
    
    def start_processing(self):
        if not self.input_path.get() or not self.output_folder.get():
            messagebox.showerror("エラー", "入力ファイルと出力フォルダーの両方のパスを選択してください。")
            return
        
        self.running = True
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.status_label.config(text="処理中...", foreground="#1E88E5")
        self.progress.start()

        # Disable the browse button when processing starts
        self.input_browse_button.config(state=tk.DISABLED)
        self.output_browse_button.config(state=tk.DISABLED)
        
        threading.Thread(target=self.process_data, daemon=True).start()
    
    def process_data(self):
        output_file_name = f"出力({get_current_time()}).xlsx"
        try:
            wb = openpyxl.load_workbook(self.input_path.get())
            ws = wb.active
            
            # Create the output workbook and sheet
            output_wb = openpyxl.Workbook()
            output_ws = output_wb.active
            output_ws.title = "処理されたデータ"
            
            # Add headers to the output Excel file
            headers = ["No", "ASIN", "商品名", "メーカー名", "販売業者", "住所", "運営責任者名", "店舗名", "URL"]
            output_ws.append(headers)
            
            total_rows = ws.max_row
            self.progress["maximum"] = total_rows
            
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if not self.running or i == 3:
                    break
                self.status_label.config(text=f"処理中 行 {i + 1}/{total_rows}...")
                self.progress["value"] = i + 1
                output_json = scrap_info(row[0])
                result = [
                    i + 1,
                    output_json.get('ASIN', ''),
                    output_json.get('商品名', ''),
                    output_json.get('メーカー名', ''),
                    output_json.get('販売業者', ''),
                    output_json.get('住所', ''),
                    output_json.get('運営責任者名', ''),
                    output_json.get('店舗名', ''),
                    output_json.get('URL', '')
                ]
                # Append the result to the output worksheet
                output_ws.append(result)
                output_file_path = os.path.join(self.output_folder.get(), output_file_name)
                output_wb.save(output_file_path)
                time.sleep(0.5)  # Simulate processing time
            
            if self.running:
                self.status_label.config(text="処理が完了しました", foreground="#388E3C")
            else:
                self.status_label.config(text="処理が停止しました", foreground="#E53935")
        except Exception as e:
            print(f"Error {str(e)}")
            messagebox.showerror("Error", str(e))
        finally:
            self.running = False
            self.start_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
            self.progress.stop()
            self.input_browse_button.config(state=tk.NORMAL)
            self.output_browse_button.config(state=tk.NORMAL)

    def stop_processing(self):
        self.running = False
        self.status_label.config(text="現在実行中のデータが完了次第、処理を停止します。", foreground="#E53935")
        self.stop_button.config(state=tk.DISABLED)

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessor(root)
    root.mainloop()
